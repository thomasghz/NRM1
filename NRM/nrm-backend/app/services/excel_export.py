"""
NRM Excel Export Service
Generates a filled .xlsx workbook for a given department + quarter
using openpyxl.
"""
import io
from datetime import date
from typing import Optional
from uuid import UUID

from sqlalchemy.orm import Session

from app.models.models import (
    Department, Risk, TreatmentPlan, KRI, KRIEntry,
    KPI, KPIEntry, ComplianceControl, ComplianceEntry, Incident,
)
from app.services.business_logic import (
    risk_zone, effective_treatment_status, indicator_status,
    risk_movement, QUARTER_LABELS, previous_quarter,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Colour constants
# ---------------------------------------------------------------------------
RED_FILL = PatternFill("solid", fgColor="FF0000") if HAS_OPENPYXL else None
AMBER_FILL = PatternFill("solid", fgColor="FFC000") if HAS_OPENPYXL else None
GREEN_FILL = PatternFill("solid", fgColor="92D050") if HAS_OPENPYXL else None
HEADER_FILL = PatternFill("solid", fgColor="1F4E79") if HAS_OPENPYXL else None
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6") if HAS_OPENPYXL else None
WHITE_FONT = Font(color="FFFFFF", bold=True) if HAS_OPENPYXL else None
BOLD = Font(bold=True) if HAS_OPENPYXL else None
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True) if HAS_OPENPYXL else None
WRAP = Alignment(wrap_text=True, vertical="top") if HAS_OPENPYXL else None


def _thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _status_fill(status: Optional[str]):
    m = {"Green": GREEN_FILL, "Amber": AMBER_FILL, "Red": RED_FILL}
    return m.get(status)


def _zone_fill(zone: Optional[str]):
    m = {
        "Low": GREEN_FILL,
        "Medium": AMBER_FILL,
        "High": PatternFill("solid", fgColor="FF6600") if HAS_OPENPYXL else None,
        "Critical": RED_FILL,
    }
    return m.get(zone)


def _header_row(ws, row: int, values: list, fill=None, font=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        cell.alignment = CENTER
        cell.border = _thin_border()


def _write_cell(ws, row: int, col: int, value, fill=None, font=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = alignment or WRAP
    cell.border = _thin_border()
    return cell


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_register(ws, dept: Department, risks: list):
    ws.title = "Register"
    ws.sheet_view.showGridLines = True

    # Department header block
    ws.merge_cells("A1:B1")
    ws["A1"] = "Department"
    ws["A1"].font = BOLD
    ws.merge_cells("C1:H1")
    ws["C1"] = dept.name

    ws.merge_cells("A2:B2")
    ws["A2"] = "Assessed Unit"
    ws["A2"].font = BOLD
    ws.merge_cells("C2:H2")
    ws["C2"] = dept.assessed_unit

    ws.merge_cells("A3:B3")
    ws["A3"] = "Objective"
    ws["A3"].font = BOLD
    ws.merge_cells("C3:H3")
    ws["C3"] = dept.objective

    # Column headers
    headers = [
        "Risk ID", "Risk Event", "CSF", "Risk Source", "Risk Effect",
        "Likelihood (I)", "Financial (I)", "Reputation (I)", "Human Capital (I)",
        "Service Delivery (I)", "Regulatory (I)", "Projects (I)", "Info Systems (I)",
        "Overall Consequence (I)", "IRL", "IRL Zone",
        "Current Controls",
        "Likelihood (R)", "Financial (R)", "Reputation (R)", "Human Capital (R)",
        "Service Delivery (R)", "Regulatory (R)", "Projects (R)", "Info Systems (R)",
        "Overall Consequence (R)", "RRL", "RRL Zone",
        "Evaluation",
    ]
    _header_row(ws, 5, headers, fill=HEADER_FILL, font=WHITE_FONT)

    for r_idx, risk in enumerate(risks, 6):
        i = 1
        _write_cell(ws, r_idx, i, str(risk.id)[:8]); i += 1
        _write_cell(ws, r_idx, i, risk.risk_event); i += 1
        _write_cell(ws, r_idx, i, risk.csf.pillar if risk.csf else ""); i += 1
        _write_cell(ws, r_idx, i, risk.risk_source); i += 1
        _write_cell(ws, r_idx, i, risk.risk_effect); i += 1
        _write_cell(ws, r_idx, i, risk.likelihood_i); i += 1
        for dim in [risk.financial_i, risk.reputation_i, risk.human_capital_i,
                    risk.service_delivery_i, risk.regulatory_i, risk.projects_i, risk.info_systems_i]:
            _write_cell(ws, r_idx, i, dim); i += 1
        _write_cell(ws, r_idx, i, risk.overall_inherent_consequence); i += 1
        irl_zone = risk_zone(risk.irl)
        _write_cell(ws, r_idx, i, risk.irl, fill=_zone_fill(irl_zone)); i += 1
        _write_cell(ws, r_idx, i, irl_zone, fill=_zone_fill(irl_zone)); i += 1
        _write_cell(ws, r_idx, i, risk.current_controls); i += 1
        _write_cell(ws, r_idx, i, risk.likelihood_r); i += 1
        for dim in [risk.financial_r, risk.reputation_r, risk.human_capital_r,
                    risk.service_delivery_r, risk.regulatory_r, risk.projects_r, risk.info_systems_r]:
            _write_cell(ws, r_idx, i, dim); i += 1
        _write_cell(ws, r_idx, i, risk.overall_residual_consequence); i += 1
        rrl_zone = risk_zone(risk.rrl)
        _write_cell(ws, r_idx, i, risk.rrl, fill=_zone_fill(rrl_zone)); i += 1
        _write_cell(ws, r_idx, i, rrl_zone, fill=_zone_fill(rrl_zone)); i += 1
        _write_cell(ws, r_idx, i, risk.evaluation); i += 1


def _build_treatment_plan(ws, risks: list, quarter: str):
    ws.title = "1.Risk Treatment Plan"
    q_label = QUARTER_LABELS.get(quarter, quarter)
    headers = [
        "Risk Event", "Improvement Action", "Due Date", "Responsibility",
        f"{q_label} — Status", f"{q_label} — Comments", f"{q_label} — Further Action",
    ]
    _header_row(ws, 1, headers, fill=HEADER_FILL, font=WHITE_FONT)

    row = 2
    q_lower = quarter.lower()
    for risk in risks:
        for plan in risk.treatment_plans:
            status = getattr(plan, f"{q_lower}_status", None)
            comments = getattr(plan, f"{q_lower}_comments", None)
            further = getattr(plan, f"{q_lower}_further_action", None)
            eff = effective_treatment_status(status, plan.due_date)
            _write_cell(ws, row, 1, risk.risk_event)
            _write_cell(ws, row, 2, plan.improvement_action)
            _write_cell(ws, row, 3, str(plan.due_date) if plan.due_date else "")
            _write_cell(ws, row, 4, plan.responsibility)
            fill = GREEN_FILL if eff == "Completed" else (RED_FILL if eff == "Overdue" else None)
            _write_cell(ws, row, 5, eff, fill=fill)
            _write_cell(ws, row, 6, comments)
            _write_cell(ws, row, 7, further)
            row += 1

    dv = DataValidation(
        type="list",
        formula1='"Not Started,WIP,Completed,Overdue"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"E2:E{row}"


def _build_kris(ws, risks: list, quarter: str):
    ws.title = "2.KRIs"
    q_label = QUARTER_LABELS.get(quarter, quarter)
    headers = [
        "Risk Event", "KRI Description", "Frequency", "Green Threshold",
        "Amber Threshold", "Direction", "Responsibility",
        f"{q_label} — Value", f"{q_label} — Status",
        "Comments", "Action Taken", "Due Date", "Responsibility (Entry)",
    ]
    _header_row(ws, 1, headers, fill=HEADER_FILL, font=WHITE_FONT)
    row = 2
    for risk in risks:
        for kri in risk.kris:
            entry = next((e for e in kri.entries if e.quarter == quarter), None)
            _write_cell(ws, row, 1, risk.risk_event)
            _write_cell(ws, row, 2, kri.description)
            _write_cell(ws, row, 3, kri.frequency)
            _write_cell(ws, row, 4, float(kri.green_threshold))
            _write_cell(ws, row, 5, float(kri.amber_threshold))
            _write_cell(ws, row, 6, kri.direction)
            _write_cell(ws, row, 7, kri.responsibility)
            if entry:
                _write_cell(ws, row, 8, float(entry.entry_value) if entry.entry_value is not None else "")
                _write_cell(ws, row, 9, entry.status, fill=_status_fill(entry.status))
                _write_cell(ws, row, 10, entry.comments)
                _write_cell(ws, row, 11, entry.action_taken)
                _write_cell(ws, row, 12, str(entry.due_date) if entry.due_date else "")
                _write_cell(ws, row, 13, entry.responsibility)
            row += 1


def _build_kpis(ws, dept: Department, quarter: str):
    ws.title = "3.KPIs"
    q_label = QUARTER_LABELS.get(quarter, quarter)
    headers = [
        "Department", "KPI Description", "Frequency",
        "Green Threshold", "Amber Threshold", "Direction", "Responsibility",
        f"{q_label} — Value", f"{q_label} — Status",
        "Comments", "Action Taken", "Due Date",
    ]
    _header_row(ws, 1, headers, fill=HEADER_FILL, font=WHITE_FONT)
    row = 2
    for kpi in dept.kpis:
        entry = next((e for e in kpi.entries if e.quarter == quarter), None)
        _write_cell(ws, row, 1, dept.name)
        _write_cell(ws, row, 2, kpi.description)
        _write_cell(ws, row, 3, kpi.frequency)
        _write_cell(ws, row, 4, float(kpi.green_threshold))
        _write_cell(ws, row, 5, float(kpi.amber_threshold))
        _write_cell(ws, row, 6, kpi.direction)
        _write_cell(ws, row, 7, kpi.responsibility)
        if entry:
            _write_cell(ws, row, 8, float(entry.entry_value) if entry.entry_value is not None else "")
            _write_cell(ws, row, 9, entry.status, fill=_status_fill(entry.status))
            _write_cell(ws, row, 10, entry.comments)
            _write_cell(ws, row, 11, entry.action_taken)
            _write_cell(ws, row, 12, str(entry.due_date) if entry.due_date else "")
        row += 1


def _build_compliance(ws, risks: list, quarter: str):
    ws.title = "4.Compliance"
    q_label = QUARTER_LABELS.get(quarter, quarter)
    headers = [
        "Risk Event", "Key Control", "Compliance Question", "Frequency", "Responsibility",
        f"{q_label} — Response", "Comments for No",
        "Action Taken", "Due Date", "Responsibility (Entry)", "Status",
    ]
    _header_row(ws, 1, headers, fill=HEADER_FILL, font=WHITE_FONT)
    row = 2
    for risk in risks:
        for ctrl in risk.compliance_controls:
            entry = next((e for e in ctrl.entries if e.quarter == quarter), None)
            _write_cell(ws, row, 1, risk.risk_event)
            _write_cell(ws, row, 2, ctrl.key_control)
            _write_cell(ws, row, 3, ctrl.compliance_question)
            _write_cell(ws, row, 4, ctrl.frequency)
            _write_cell(ws, row, 5, ctrl.responsibility)
            if entry:
                resp_fill = GREEN_FILL if entry.response == "Yes" else (RED_FILL if entry.response == "No" else None)
                _write_cell(ws, row, 6, entry.response, fill=resp_fill)
                _write_cell(ws, row, 7, entry.comments_for_no)
                _write_cell(ws, row, 8, entry.action_taken)
                _write_cell(ws, row, 9, str(entry.due_date) if entry.due_date else "")
                _write_cell(ws, row, 10, entry.responsibility)
                _write_cell(ws, row, 11, entry.status)
            row += 1

    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.sqref = f"F2:F{row}"


def _build_incidents(ws, incidents: list):
    ws.title = "5.Incidents Management"
    headers = [
        "Serial No", "Quarter", "Details", "Incident Date", "Location",
        "Direct Financial Impact", "Non-Financial Impact", "Action Taken",
        "Risk Causes", "Risk Effects", "Control Failed", "Corrective Action",
        "Case Summary Done", "Manager Comments", "Further Corrective Action", "Due Date",
    ]
    _header_row(ws, 1, headers, fill=HEADER_FILL, font=WHITE_FONT)
    for r_idx, inc in enumerate(incidents, 2):
        vals = [
            inc.serial_no, inc.quarter, inc.details,
            str(inc.incident_date) if inc.incident_date else "",
            inc.location,
            "Yes" if inc.direct_financial_impact else "No",
            inc.non_financial_impact, inc.action_taken,
            inc.risk_causes, inc.risk_effects, inc.control_failed,
            inc.corrective_action,
            "Yes" if inc.case_summary_done else "No",
            inc.managers_comments, inc.further_corrective_action,
            str(inc.due_date) if inc.due_date else "",
        ]
        for c_idx, val in enumerate(vals, 1):
            _write_cell(ws, r_idx, c_idx, val)


def _build_aggregation(ws, risks: list, quarter: str):
    ws.title = "Risk Aggregation"
    q_label = QUARTER_LABELS.get(quarter, quarter)
    prev_q = previous_quarter(quarter)
    headers = [
        "Rank", "Risk Event", "Movement", "IRL", "IRL Zone", "RRL", "RRL Zone",
        "Latest KRI Value", "KRI Status", "KRI Corrective Action Status",
        "Compliance Response", "Compliance Corrective Action Status",
        "# Incidents", "Treatment Status",
        "Risk Owner Comments", "Action Being Taken", "Due Date", "Responsible Person",
    ]
    _header_row(ws, 1, headers, fill=HEADER_FILL, font=WHITE_FONT)

    # Gather aggregation override rows indexed by risk_id
    agg_map = {}
    for risk in risks:
        for agg in risk.aggregations:
            if agg.quarter == quarter:
                agg_map[risk.id] = agg

    row = 2
    sorted_risks = sorted(
        risks,
        key=lambda r: agg_map[r.id].aggregate_ranking
        if r.id in agg_map and agg_map[r.id].aggregate_ranking
        else 999,
    )

    for risk in sorted_risks:
        agg = agg_map.get(risk.id)
        rank = agg.aggregate_ranking if agg else None

        # KRI: latest entry for this quarter
        latest_kri_val = None
        latest_kri_status = None
        kri_action_status = None
        for kri in risk.kris:
            entry = next((e for e in kri.entries if e.quarter == quarter), None)
            if entry:
                latest_kri_val = float(entry.entry_value) if entry.entry_value is not None else None
                latest_kri_status = entry.status
                kri_action_status = entry.previous_action_status

        # Compliance
        comp_response = None
        comp_action_status = None
        for ctrl in risk.compliance_controls:
            entry = next((e for e in ctrl.entries if e.quarter == quarter), None)
            if entry:
                comp_response = entry.response
                comp_action_status = entry.status

        # Incidents
        inc_count = sum(1 for i in risk.incidents if i.quarter == quarter)

        # Treatment
        treatment_status = None
        q_lower = quarter.lower()
        for plan in risk.treatment_plans:
            s = getattr(plan, f"{q_lower}_status", None)
            treatment_status = effective_treatment_status(s, plan.due_date)
            break

        # Movement
        prev_rrl = None
        if prev_q:
            for agg2 in risk.aggregations:
                if agg2.quarter == prev_q:
                    prev_rrl = risk.rrl  # simplified; ideally store historical RRL
        movement = risk_movement(risk.rrl, prev_rrl) if prev_rrl else "\u2192"

        irl_zone = risk_zone(risk.irl)
        rrl_zone = risk_zone(risk.rrl)

        vals = [
            rank, risk.risk_event, movement,
            risk.irl, irl_zone, risk.rrl, rrl_zone,
            latest_kri_val, latest_kri_status, kri_action_status,
            comp_response, comp_action_status, inc_count, treatment_status,
            agg.risk_owner_comments if agg else None,
            agg.action_being_taken if agg else None,
            str(agg.due_date) if agg and agg.due_date else None,
            agg.responsible_person if agg else None,
        ]
        fills = [
            None, None, None,
            _zone_fill(irl_zone), _zone_fill(irl_zone),
            _zone_fill(rrl_zone), _zone_fill(rrl_zone),
            None, _status_fill(latest_kri_status), None,
            GREEN_FILL if comp_response == "Yes" else (RED_FILL if comp_response == "No" else None),
            None, None,
            GREEN_FILL if treatment_status == "Completed" else (RED_FILL if treatment_status == "Overdue" else None),
            None, None, None, None,
        ]
        for c_idx, (val, fill) in enumerate(zip(vals, fills), 1):
            _write_cell(ws, row, c_idx, val, fill=fill)
        row += 1


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def generate_workbook(db: Session, department_id: UUID, quarter: str) -> bytes:
    """
    Generate a fully populated NRM .xlsx workbook and return it as bytes.
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    dept = (
        db.query(Department)
        .filter(Department.id == department_id)
        .first()
    )
    if not dept:
        raise ValueError(f"Department {department_id} not found")

    from sqlalchemy.orm import joinedload
    risks = (
        db.query(Risk)
        .options(
            joinedload(Risk.csf),
            joinedload(Risk.treatment_plans),
            joinedload(Risk.kris).joinedload(KRI.entries),
            joinedload(Risk.compliance_controls).joinedload(ComplianceControl.entries),
            joinedload(Risk.incidents),
            joinedload(Risk.aggregations),
        )
        .filter(Risk.department_id == department_id)
        .all()
    )

    dept_kpis = (
        db.query(KPI)
        .options(joinedload(KPI.entries))
        .filter(KPI.department_id == department_id)
        .all()
    )
    dept.kpis = dept_kpis  # attach for sheet builder

    incidents_all = (
        db.query(Incident)
        .filter(Incident.department_id == department_id)
        .all()
    )

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_reg = wb.create_sheet("Register")
    _build_register(ws_reg, dept, risks)

    ws_tp = wb.create_sheet("1.Risk Treatment Plan")
    _build_treatment_plan(ws_tp, risks, quarter)

    ws_kri = wb.create_sheet("2.KRIs")
    _build_kris(ws_kri, risks, quarter)

    ws_kpi = wb.create_sheet("3.KPIs")
    _build_kpis(ws_kpi, dept, quarter)

    ws_comp = wb.create_sheet("4.Compliance")
    _build_compliance(ws_comp, risks, quarter)

    ws_inc = wb.create_sheet("5.Incidents Management")
    _build_incidents(ws_inc, incidents_all)

    ws_agg = wb.create_sheet("Risk Aggregation")
    _build_aggregation(ws_agg, risks, quarter)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
